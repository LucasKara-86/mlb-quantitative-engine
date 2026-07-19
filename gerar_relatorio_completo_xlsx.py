"""Gera o relatório completo do dia em XLSX, incluindo Value Bet de Game Total e
Team Total (por time), reaproveitando as projeções já calculadas nesta sessão
(lidas de relatorio_2026-07-18.csv) e buscando apenas dados que mudam rápido:
lineup (confiança), odds de jogo e odds de team_totals (via endpoint por evento).
"""

import time

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mlb_quantitative_engine.analytics.value_bet_calculator import (
    evaluate_game_total_value_bets,
    evaluate_team_total_value_bets,
)
from mlb_quantitative_engine.api.mlb_api import MLBApiClient
from mlb_quantitative_engine.database.repository import Repository
from mlb_quantitative_engine.services.lineup_service import LineupService
from mlb_quantitative_engine.services.odds_service import OddsService

DATE = "2026-07-18"
CACHED_CSV = "relatorio_2026-07-18.csv"
OUTPUT_XLSX = f"relatorio_{DATE}_completo.xlsx"

client = MLBApiClient()
lineup_service = LineupService(client)
odds_service = OddsService()
repo = Repository(db_path="mlb_quantitative_engine/database/database.db")

cached = pd.read_csv(CACHED_CSV, sep=None, engine="python", encoding="utf-8-sig")
cached = cached[cached["skip_reason"].isna()].copy()
print(f"Jogos com projeção válida (cache): {len(cached)}", flush=True)

print("Buscando odds de jogo atualizadas...", flush=True)
all_game_odds = odds_service.get_all_game_odds()
print(f"Odds de jogo: {len(all_game_odds)} eventos", flush=True)

rows = []
for i, row in enumerate(cached.itertuples(), 1):
    t0 = time.time()
    game_pk = row.game_pk
    home_team, away_team = row.home_team, row.away_team

    # 1. Confiança atualizada (lineup pode já estar oficial agora)
    home_lineup = lineup_service.get_batting_order(game_pk, "home")
    away_lineup = lineup_service.get_batting_order(game_pk, "away")
    confidence_score = (home_lineup.confidence_score + away_lineup.confidence_score) / 2.0

    # 2. Odds de jogo (Game Total) já atualizadas
    game_odds = odds_service.find_game_odds(all_game_odds, home_team, away_team)
    consensus = game_odds.consensus_total if game_odds else None

    candidates = []

    if consensus is not None:
        game_over, game_under = evaluate_game_total_value_bets(
            game_pk=game_pk, home_team=home_team, away_team=away_team,
            projected_total_runs=row.projected_total_runs, point=consensus.point,
            over_price=consensus.over_price, over_bookmaker=consensus.over_bookmaker,
            under_price=consensus.under_price, under_bookmaker=consensus.under_bookmaker,
            confidence_score=confidence_score,
        )
        candidates.extend([game_over, game_under])

    # 3. Odds de Team Total (endpoint por evento -- 1 crédito por jogo)
    if game_odds is not None and game_odds.event_id:
        team_totals = odds_service.get_team_totals(game_odds.event_id, home_team, away_team)

        if team_totals.home is not None:
            home_over, home_under = evaluate_team_total_value_bets(
                game_pk=game_pk, home_team=home_team, away_team=away_team,
                team_label="home_team_total", projected_team_runs=row.projected_home_runs,
                point=team_totals.home.point, over_price=team_totals.home.over_price,
                over_bookmaker=team_totals.home.over_bookmaker, under_price=team_totals.home.under_price,
                under_bookmaker=team_totals.home.under_bookmaker, confidence_score=confidence_score,
            )
            candidates.extend([home_over, home_under])

        if team_totals.away is not None:
            away_over, away_under = evaluate_team_total_value_bets(
                game_pk=game_pk, home_team=home_team, away_team=away_team,
                team_label="away_team_total", projected_team_runs=row.projected_away_runs,
                point=team_totals.away.point, over_price=team_totals.away.over_price,
                over_bookmaker=team_totals.away.over_bookmaker, under_price=team_totals.away.under_price,
                under_bookmaker=team_totals.away.under_bookmaker, confidence_score=confidence_score,
            )
            candidates.extend([away_over, away_under])

    # 4. Persiste todas as avaliações no banco (histórico completo)
    game_row = repo.get_game_by_pk(game_pk)
    projections = repo.list_projections_for_game(game_row.id) if game_row else []
    projection_id = projections[0].id if projections else None
    if projection_id:
        for bet in candidates:
            repo.save_value_bet(
                projection_id=projection_id, market=bet.market, bookmaker=bet.bookmaker,
                price=bet.price, point=bet.point, projection_probability=bet.projected_probability,
                implied_probability_raw=bet.implied_probability_raw, implied_probability_fair=bet.implied_probability_fair,
                edge=bet.edge, expected_value=bet.expected_value, kelly_fraction=bet.kelly_fraction,
                kelly_fraction_quarter=bet.kelly_fraction_quarter, confidence_score=bet.confidence_score,
                meets_criteria=bet.meets_criteria,
            )

    # 5. Melhor aposta do jogo (para o resumo)
    qualifying = [b for b in candidates if b.meets_criteria]
    best = max(qualifying, key=lambda b: b.expected_value) if qualifying else (
        max(candidates, key=lambda b: b.expected_value) if candidates else None
    )

    market_labels = {
        "game_total_over": "Jogo OVER", "game_total_under": "Jogo UNDER",
        "home_team_total_over": f"{home_team} OVER", "home_team_total_under": f"{home_team} UNDER",
        "away_team_total_over": f"{away_team} OVER", "away_team_total_under": f"{away_team} UNDER",
    }

    rows.append({
        "jogo": f"{away_team} @ {home_team}",
        "horario": row.game_datetime,
        "pitchers": f"{row.away_probable_pitcher} vs {row.home_probable_pitcher}",
        "projecao_casa": row.projected_home_runs,
        "projecao_visitante": row.projected_away_runs,
        "projecao_total": row.projected_total_runs,
        "confianca": round(confidence_score, 1),
        "melhor_aposta": market_labels.get(best.market, best.market) if best else None,
        "linha": best.point if best else None,
        "odd": best.price if best else None,
        "casa_apostas": best.bookmaker if best else None,
        "edge": best.edge if best else None,
        "ev": best.expected_value if best else None,
        "kelly_1_4": best.kelly_fraction_quarter if best else None,
        "recomendado": bool(best.meets_criteria) if best else False,
    })
    print(f"[{i}/{len(cached)}] {away_team} @ {home_team} -> "
          f"melhor={rows[-1]['melhor_aposta']} ev={rows[-1]['ev']} rec={rows[-1]['recomendado']} "
          f"({round(time.time()-t0,1)}s)", flush=True)

df = pd.DataFrame(rows)
df["ev"] = df["ev"].fillna(-999)
df = df.sort_values(by=["recomendado", "ev"], ascending=[False, False]).drop(columns=[]).reset_index(drop=True)
df.loc[df["ev"] == -999, "ev"] = None

# --- Exportação em xlsx com formatação ---
try:
    df.to_excel(OUTPUT_XLSX, index=False, sheet_name="Relatorio")
except PermissionError:
    OUTPUT_XLSX = OUTPUT_XLSX.replace(".xlsx", "_v2.xlsx")
    print(f"Arquivo original bloqueado (aberto em outro programa?). Salvando como {OUTPUT_XLSX}", flush=True)
    df.to_excel(OUTPUT_XLSX, index=False, sheet_name="Relatorio")

wb = load_workbook(OUTPUT_XLSX)
ws = wb["Relatorio"]

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
even_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
recommended_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

recomendado_col = df.columns.get_loc("recomendado") + 1
for row_idx in range(2, ws.max_row + 1):
    is_recommended = ws.cell(row=row_idx, column=recomendado_col).value
    fill = recommended_fill if is_recommended else (even_fill if row_idx % 2 == 0 else odd_fill)
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill

for col_idx, column in enumerate(df.columns, 1):
    max_len = max(len(str(column)), df[column].astype(str).map(len).max() if len(df) else 0)
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

ws.freeze_panes = "A2"
wb.save(OUTPUT_XLSX)

print(f"\nOK -> {OUTPUT_XLSX}", flush=True)
print(f"Creditos odds usados: {odds_service.api_client.last_requests_used}, "
      f"restantes: {odds_service.api_client.last_requests_remaining}", flush=True)
