from __future__ import annotations

"""Exporta o estado atual do banco (jogos + projeção mais recente + melhor Value Bet
de cada jogo) para um arquivo xlsx formatado.

O banco é a fonte da verdade (ver ReportGenerator) — este módulo só lê o que já
foi persistido e nunca busca dados novos. Pode ser chamado a qualquer momento
para obter um snapshot atualizado do relatório, inclusive várias vezes ao longo
do dia conforme lotes de jogos vão sendo processados incrementalmente
(ver analytics/batch_scheduling.py e reports/incremental_runner.py).
"""

from dataclasses import asdict, dataclass
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mlb_quantitative_engine.database.repository import Repository
from mlb_quantitative_engine.models.value_bet import describe_market


@dataclass(frozen=True)
class ExportRow:
    """Uma linha do relatório exportado: um jogo com sua projeção e melhor Value Bet."""

    game_pk: int
    jogo: str
    horario: Optional[str]
    pitchers: str
    projecao_casa: Optional[float]
    projecao_visitante: Optional[float]
    projecao_total: Optional[float]
    confianca: Optional[float]
    melhor_aposta: Optional[str]
    linha: Optional[float]
    odd_minima: Optional[float]
    edge: Optional[float]
    ev: Optional[float]
    stake_sugerida: Optional[float]
    recomendado: bool


def build_export_rows(repository: Repository, game_date: str) -> List[ExportRow]:
    """Monta as linhas do relatório a partir do estado atual do banco para uma data."""
    games = repository.list_games_by_date(game_date)
    rows: List[ExportRow] = []

    for game in games:
        projections = repository.list_projections_for_game(game.id)
        projection = projections[0] if projections else None
        bets = repository.list_value_bets_for_projection(projection.id) if projection else []

        qualifying = [bet for bet in bets if bet.meets_criteria]
        best = max(qualifying, key=lambda bet: bet.expected_value) if qualifying else (
            max(bets, key=lambda bet: bet.expected_value) if bets else None
        )

        rows.append(
            ExportRow(
                game_pk=game.game_pk,
                jogo=f"{game.away_team} @ {game.home_team}",
                horario=game.game_datetime,
                pitchers=f"{game.away_probable_pitcher} vs {game.home_probable_pitcher}",
                projecao_casa=projection.projected_home_runs if projection else None,
                projecao_visitante=projection.projected_away_runs if projection else None,
                projecao_total=projection.projected_total_runs if projection else None,
                confianca=best.confidence_score if best else None,
                melhor_aposta=describe_market(best.market, game.home_team, game.away_team) if best else None,
                linha=best.point if best else None,
                odd_minima=best.minimum_acceptable_price if best else None,
                edge=best.edge if best else None,
                ev=best.expected_value if best else None,
                stake_sugerida=best.suggested_stake_fraction if best else None,
                recomendado=bool(best.meets_criteria) if best else False,
            )
        )

    return rows


def export_to_xlsx(repository: Repository, game_date: str, output_path: str) -> str:
    """Gera o xlsx formatado (cabeçalho colorido, linhas zebradas, recomendados em
    verde, ordenado da melhor para a pior aposta). Retorna o caminho efetivamente
    salvo (pode diferir de `output_path` se o arquivo estiver bloqueado)."""
    rows = build_export_rows(repository, game_date)
    df = pd.DataFrame([asdict(row) for row in rows])

    if df.empty:
        df = pd.DataFrame(columns=[f.name for f in ExportRow.__dataclass_fields__.values()])
    else:
        sort_key = pd.to_numeric(df["ev"], errors="coerce").fillna(-999.0)
        df = df.assign(_sort_key=sort_key).sort_values(
            by=["recomendado", "_sort_key"], ascending=[False, False]
        ).drop(columns=["_sort_key"]).reset_index(drop=True)

    try:
        df.to_excel(output_path, index=False, sheet_name="Relatorio")
    except PermissionError:
        output_path = output_path.replace(".xlsx", "_v2.xlsx")
        df.to_excel(output_path, index=False, sheet_name="Relatorio")

    _apply_formatting(output_path, df)
    return output_path


def _apply_formatting(output_path: str, df: pd.DataFrame) -> None:
    wb = load_workbook(output_path)
    ws = wb["Relatorio"]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    even_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    recommended_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    if "recomendado" in df.columns:
        recomendado_col = df.columns.get_loc("recomendado") + 1
        for row_idx in range(2, ws.max_row + 1):
            is_recommended = ws.cell(row=row_idx, column=recomendado_col).value
            fill = recommended_fill if is_recommended else (even_fill if row_idx % 2 == 0 else odd_fill)
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    for col_idx, column in enumerate(df.columns, 1):
        max_len = max(len(str(column)), df[column].astype(str).map(len).max() if len(df) else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"
    wb.save(output_path)
